import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import simpledialog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from copy import copy
from tkcalendar import DateEntry
from copy import copy

# Read data from Excel file
excel_file_path = r"C:\Users\vinay.jha\OneDrive - UKG\MS Garnishment Services\Noida Team\POD_Utilization\POD Utilization1.xlsx"
df = pd.read_excel(excel_file_path, sheet_name="Sheet2")
ws = pd.read_excel(excel_file_path, sheet_name="Utilisation")

# Create the main window
root = tk.Tk()
root.title('POD Utilization Form')

# Create a frame to hold the widgets
frame = ttk.Frame(root, padding="10")
frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

# Define the widgets
pod_name_label = ttk.Label(frame, text='Pod Name')
pod_name = ttk.Combobox(frame, values=df['POD'].dropna().unique().tolist())
pod_members_label = ttk.Label(frame, text='Pod Members')
pod_members = tk.Listbox(frame, selectmode='multiple')
activities_label = ttk.Label(frame, text='Activities')
activities = ttk.Combobox(frame)
start_time_label = ttk.Label(frame, text='Start Time')
start_time = ttk.Entry(frame)
end_time_label = ttk.Label(frame, text='End Time')
end_time = ttk.Entry(frame)
date_label = ttk.Label(frame, text='Date')
date = DateEntry(frame)  # Use DateEntry widget for date selection
comments_label = ttk.Label(frame, text='Comments')
comments = tk.Text(frame, width=25, height=5)

# Function to open a new form for time selection
def open_time_selection(entry_field):
    def select_time():
        selected_time = f"{hours.get()}:{minutes.get()}"  # Remove the space after the colon
        entry_field.delete(0, tk.END)
        entry_field.insert(0, selected_time)
        time_selection_window.destroy()

    time_selection_window = tk.Toplevel(root)
    time_selection_window.title('Select Time')

    hours_label = ttk.Label(time_selection_window, text='Hours')
    hours = ttk.Combobox(time_selection_window, values=[f'{i:02d}' for i in range(24)])
    minutes_label = ttk.Label(time_selection_window, text='Minutes')
    minutes = ttk.Combobox(time_selection_window, values=[f'{i:02d}' for i in range(60)])
    select_button = ttk.Button(time_selection_window, text='Select', command=select_time)

    hours_label.grid(row=0, column=0, sticky=tk.W)
    hours.grid(row=0, column=1, sticky=(tk.W, tk.E))
    minutes_label.grid(row=1, column=0, sticky=tk.W)
    minutes.grid(row=1, column=1, sticky=(tk.W, tk.E))
    select_button.grid(row=2, column=0, columnspan=2)

# Create buttons for time selection
start_time_button = ttk.Button(frame, text='Select Time', command=lambda: open_time_selection(start_time))
end_time_button = ttk.Button(frame, text='Select Time', command=lambda: open_time_selection(end_time))

def update_pod_members_and_activities(event):
    print("update_pod_members_and_activities function called")
    value = pod_name.get()
    if value == 'POD_1':
        members = df['POD_1'].dropna().tolist()
        activities_list = df['Activities'].dropna().tolist()
    elif value == 'POD_2':
        members = df['POD_2'].dropna().tolist()
        activities_list = df['Activities'].dropna().tolist()
    elif value == 'POD_3':
        members = df['POD_3'].dropna().tolist()
        activities_list = df['Activities'].dropna().tolist()
    elif value == 'POD_4':
        members = df['POD_4'].dropna().tolist()
        activities_list = df['Activities'].dropna().tolist()
    elif value == 'Business Analyst':
        members = df['Business Analyst'].dropna().tolist()
        activities_list = df['Business Analyst Activities'].dropna().tolist()
    elif value == 'Process Leads':
        members = df['Process Leads'].dropna().tolist()
        activities_list = df['Process Leads Activities'].dropna().tolist()
    elif value == 'Implementation Analyst':
        members = df['Implementation Analyst'].dropna().tolist()
        activities_list = df['Implementation Analyst Activities'].dropna().tolist()
    else:
        members = []
        activities_list = []

    # Clear the Listbox widget before adding new members
    pod_members.delete(0, tk.END)

    # Add each member to the Listbox widget
    for member in members:
         pod_members.insert(tk.END, member)
    
    # Update the activities dropdown
    activities['values'] = activities_list

# Bind the function to the 'pod_name' dropdown
pod_name.bind('<<ComboboxSelected>>', update_pod_members_and_activities)

def copy_styles_from_row(ws, source_row, target_row):
    for col_num in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_num)
        source_cell = ws['{}{}'.format(col_letter, source_row)]
        target_cell = ws['{}{}'.format(col_letter, target_row)]
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

# Function to save data to the Utilisation worksheet and format the row
def save_to_excel(data):
    print(f"Saving the following data to {excel_file_path}: {data}")
    wb = load_workbook(excel_file_path)
    ws = wb['Utilisation']
    
    for entry in data:
        ws.append(entry)
    
        # Copy the format from the second row to the newly appended row
        new_row = ws.max_row
        copy_styles_from_row(ws, 2, new_row)
    
    wb.save(excel_file_path)  # Save the workbook after appending the data
    print("Data saved successfully")

# Function to calculate the difference between end time and start time in minutes
def calculate_time_difference(start_time_str, end_time_str):
    start_time = datetime.strptime(start_time_str, '%H:%M')
    end_time = datetime.strptime(end_time_str, '%H:%M')
    time_difference = end_time - start_time
    minutes_difference = time_difference.total_seconds() / 60
    return minutes_difference

# Function to convert date and time from string to datetime format
def convert_to_datetime(date_obj, time_str):
    date_format = '%Y-%m-%d'  # Define the date format
    time_format = '%H:%M'     # Define the time format
    
    # Convert date and time strings to datetime objects
    date_str = date_obj.strftime(date_format)
    date_obj = datetime.strptime(date_str, date_format).date()
    time_obj = datetime.strptime(time_str, time_format).time()
    
    return date_obj, time_obj

# Function to handle form submission
def submit():
    pod_name_value = pod_name.get()
    print(pod_name_value)
    pod_members_values = [pod_members.get(i) for i in pod_members.curselection()]
    if not pod_members_values:
        messagebox.showerror('Error', 'Please select at least one member')
        return
    print(pod_members_values)
    activity_value = activities.get()
    print(activity_value)
    start_time_str = start_time.get()
    print(start_time_str)
    end_time_str = end_time.get()
    print(end_time_str)
    date_obj = date.get_date()  # Use get_date method to get the selected date
    print(date_obj)
    comments_value = comments.get('1.0', tk.END)
    print(comments_value)

    # Convert date and start time to datetime format
    date_obj, start_time_obj = convert_to_datetime(date_obj, start_time_str)
    
    # Convert end time to time format
    _, end_time_obj = convert_to_datetime(date_obj, end_time_str)  # Date is not used for end time
    
    # Calculate the difference in minutes between start time and end time
    hours = calculate_time_difference(start_time_str, end_time_str)
    
    # Prepare data for each selected POD member
    data_to_save = []
    for member in pod_members_values:
        # Append the data in the order of the columns in the Excel file
        data_to_save.append([pod_name_value, member, activity_value, date_obj, start_time_obj, end_time_obj, hours, comments_value])
    
    # Save data to Excel
    save_to_excel(data_to_save)
    
    # Show popup message
    messagebox.showinfo('Success', 'Data saved successfully')
    
    # Reset form fields
    pod_name.set('')
    pod_members.delete(0, tk.END)
    activities.set('')
    start_time.delete(0, tk.END)  # Clear the Entry field
    end_time.delete(0, tk.END)  # Clear the Entry field
    date.set_date(datetime.now())  # Reset the date to today's date
    comments.delete('1.0', tk.END)
        
# Function to handle form cancellation
def cancel():
    root.destroy()

# Create the buttons
submit_button = ttk.Button(frame, text='Submit', command=submit)
cancel_button = ttk.Button(frame, text='Cancel', command=cancel)

# Add the widgets to the frame
pod_name_label.grid(row=0, column=0, sticky=tk.W)
pod_name.grid(row=0, column=1, sticky=(tk.W, tk.E))
pod_members_label.grid(row=1, column=0, sticky=tk.W)
pod_members.grid(row=1, column=1, sticky=(tk.W, tk.E))
activities_label.grid(row=2, column=0, sticky=tk.W)
activities.grid(row=2, column=1, sticky=(tk.W, tk.E))
start_time_label.grid(row=3, column=0, sticky=tk.W)
start_time.grid(row=3, column=1, sticky=(tk.W, tk.E))
start_time_button.grid(row=3, column=2, sticky=tk.W)  # Add the start time button
end_time_label.grid(row=4, column=0, sticky=tk.W)
end_time.grid(row=4, column=1, sticky=(tk.W, tk.E))
end_time_button.grid(row=4, column=2, sticky=tk.W)  # Add the end time button
date_label.grid(row=5, column=0, sticky=tk.W)
date.grid(row=5, column=1, sticky=(tk.W, tk.E))
comments_label.grid(row=6, column=0, sticky=tk.W)
comments.grid(row=6, column=1, sticky=(tk.W, tk.E))
submit_button.grid(row=7, column=0, sticky=tk.W)
cancel_button.grid(row=7, column=1, sticky=tk.E)

# Make the frame expand with the window
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)
frame.columnconfigure(1, weight=1)

# Start the main event loop
root.mainloop()